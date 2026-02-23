#!/usr/bin/env python3
from pathlib import Path
import re
from openpyxl import Workbook

MD_PATH = Path('Doc/Testing/system_test_specification.md')
XLSX_PATH = Path('Doc/Testing/system_test_specification_tables.xlsx')
SELECT_LINES = 67
SECTION_KEYS = {
    '3.1': '3.1 一般ユーザーシナリオ',
    '3.2': '3.2 管理ユーザーシナリオ',
    '3.3': '3.3 エラーハンドリングシナリオ',
}
FORBIDDEN_SHEET = r'[:\\/?*\[\]]'

import re

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(FORBIDDEN_SHEET, '_', name)
    return name.strip()[:31]


def split_tab_row(r: str):
    return [c.strip() for c in r.split('\t')]


def parse_converted_md(lines):
    selection = lines[:SELECT_LINES]
    tables_by_sub = {k: [] for k in SECTION_KEYS.keys()}
    individual = {}
    current_sub = None
    current_test = None
    i = 0
    while i < len(lines):
        line = lines[i]
        m = re.match(r'^(#{1,6})\s*(.*)$', line)
        if m:
            heading = m.group(2).strip()
            for key in SECTION_KEYS.keys():
                if heading.startswith(key):
                    current_sub = key
                    break
            else:
                tm = re.match(r'^(UST|AST|EST)-\d{3}[:_\s-]*(.*)$', heading)
                if tm:
                    code_match = re.match(r'^(UST|AST|EST)-\d{3}', heading)
                    if code_match:
                        code_id = code_match.group(0)
                        title = heading[len(code_id):].strip(' :-_')
                        sheet_name = f"{code_id}_{title}" if title else code_id
                        sheet_name = sanitize_sheet_name(sheet_name)
                        current_test = sheet_name
                        individual[current_test] = []
                    else:
                        current_test = None
                else:
                    current_test = None
            i += 1
            continue
        # detect tab table block: header contains tab and following line also contains tab
        if '\t' in line:
            header = split_tab_row(line)
            j = i+1
            rows = []
            while j < len(lines) and '\t' in lines[j]:
                rows.append(split_tab_row(lines[j]))
                j += 1
            # store
            if current_test is not None:
                individual[current_test].append((header, rows))
            elif current_sub in tables_by_sub:
                tables_by_sub[current_sub].append((header, rows))
            i = j
            continue
        i += 1
    return selection, tables_by_sub, individual


def create_workbook(selection, tables_by_sub, individual, out_path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    ws_sel = wb.create_sheet(title='選択範囲')
    ws_sel.append(['content'])
    for ln in selection:
        ws_sel.append([ln])
    for key, title in SECTION_KEYS.items():
        ws = wb.create_sheet(title=title[:31])
        ws.append([title])
        ws.append([''])
        tables = tables_by_sub.get(key, [])
        if not tables:
            ws.append(['(No tables found in this section)'])
            continue
        for idx, (header, rows) in enumerate(tables, start=1):
            ws.append([f'Table {idx}'])
            ws.append(header)
            for r in rows:
                ws.append(r)
            ws.append([''])
    for sheet, tables in individual.items():
        ws = wb.create_sheet(title=sheet)
        if not tables:
            ws.append(['(No tables found)'])
            continue
        for idx, (header, rows) in enumerate(tables, start=1):
            ws.append([f'Table {idx}'])
            ws.append(header)
            for r in rows:
                ws.append(r)
            ws.append([''])
    wb.save(out_path)
    print('Recreated workbook with', len(tables_by_sub.get('3.1',[])), 'tables in 3.1 and', len(individual), 'individual tests')

if __name__ == '__main__':
    if not MD_PATH.exists():
        print('md missing'); raise SystemExit(1)
    text = MD_PATH.read_text(encoding='utf-8')
    lines = text.splitlines()
    selection, tables_by_sub, individual = parse_converted_md(lines)
    create_workbook(selection, tables_by_sub, individual, XLSX_PATH)
