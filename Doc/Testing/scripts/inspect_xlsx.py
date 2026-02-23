from openpyxl import load_workbook
from pathlib import Path
wb = load_workbook(Path('Doc/Testing/system_test_specification_tables.xlsx'))
print('SHEET_NAMES:')
for s in wb.sheetnames:
    print('-', s)
    ws = wb[s]
    # print first 5 rows
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        print('  ', row)
        if i >= 5:
            break
    print()
