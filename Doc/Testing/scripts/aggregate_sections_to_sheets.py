#!/usr/bin/env python3
from pathlib import Path
import re
from openpyxl import load_workbook


def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', '_', name)
    name = name.strip()
    if len(name) == 0:
        return 'Sheet'
    return name[:31]


TARGET = {
    '4': '4. 一般ユーザーシナリオ',
    '5': '5. 管理ユーザーシナリオ',
    '6': '6. エラーハンドリングシナリオ',
}


def parse_tables_by_section(lines):
    tables_by_section = {k: [] for k in TARGET.keys()}
    last_heading = None
    current_section = None
    i = 0
    while i < len(lines):
        line = lines[i]
        # detect section headings
        m = re.match(r'^(#{1,6})\s*(.*)$', line)
        if m:
            last_heading = m.group(2).strip()
            # check if it's one of target sections
            for key, title in TARGET.items():
                if last_heading.startswith(title):
                    current_section = key
                    break
            else:
                # not a target section
                if line.startswith('## '):
                    current_section = None
            i += 1
            continue

        if '|' in line:
            # lookahead for separator line
            if i+1 < len(lines) and re.search(r'^[\s\|:\-]+$', lines[i+1].replace(' ', '')):
                header_line = line
                sep_line = lines[i+1]
                rows = []
                i += 2
                while i < len(lines) and '|' in lines[i]:
                    if re.match(r'^\s*$', lines[i]):
                        break
                    rows.append(lines[i])
                    i += 1

                def split_row(r):
                    r = r.strip()
                    if r.startswith('|') and r.endswith('|'):
                        r = r[1:-1]
                    parts = [c.strip() for c in r.split('|')]
                    return parts

                header_cells = split_row(header_line)
                data = [split_row(r) for r in rows if r.strip()]
                # store under current_section if present
                if current_section in tables_by_section:
                    tables_by_section[current_section].append((header_cells, data))
                # else ignore or could store globally
                continue
        i += 1
    return tables_by_section


def write_section_sheets(md_path: Path, xlsx_path: Path):
    text = md_path.read_text(encoding='utf-8')
    lines = text.splitlines()
    tables_by_section = parse_tables_by_section(lines)

    wb = load_workbook(xlsx_path)

    for key, title in TARGET.items():
        sheet_title = sanitize_sheet_name(title)
        # remove existing sheet if present
        if sheet_title in wb.sheetnames:
            wb.remove(wb[sheet_title])
        ws = wb.create_sheet(title=sheet_title)
        # write a simple header
        ws.append([title])
        ws.append([''])
        tables = tables_by_section.get(key, [])
        if not tables:
            ws.append(['(No tables found in this section)'])
            continue
        for idx, (header, rows) in enumerate(tables, start=1):
            # table label
            ws.append([f'Table {idx}'])
            # header row
            ws.append(header)
            # data rows
            for r in rows:
                ws.append(r)
            # spacer
            ws.append([''])

    wb.save(xlsx_path)
    print(f'Wrote section sheets for {list(TARGET.values())} into {xlsx_path}')


if __name__ == '__main__':
    import sys
    if len(sys.argv) != 3:
        print('Usage: aggregate_sections_to_sheets.py input.md output.xlsx')
        raise SystemExit(1)
    md = Path(sys.argv[1])
    out = Path(sys.argv[2])
    if not md.exists():
        print('Input md not found:', md)
        raise SystemExit(1)
    if not out.exists():
        print('Output xlsx not found:', out)
        raise SystemExit(1)
    write_section_sheets(md, out)
