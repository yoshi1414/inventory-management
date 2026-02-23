from pathlib import Path
from openpyxl import load_workbook

MD_PATH = Path('Doc/Testing/system_test_specification.md')
XLSX_PATH = Path('Doc/Testing/system_test_specification_tables.xlsx')
SHEET_NAME = '選択範囲'
LINE_COUNT = 67  # 選択範囲はファイル先頭の1-67行

if not MD_PATH.exists():
    print(f"Markdown file not found: {MD_PATH}")
    raise SystemExit(1)
if not XLSX_PATH.exists():
    print(f"XLSX file not found: {XLSX_PATH}")
    raise SystemExit(1)

text = MD_PATH.read_text(encoding='utf-8')
lines = text.splitlines()
selected = lines[:LINE_COUNT]

wb = load_workbook(XLSX_PATH)
# Remove existing sheet with same name to avoid duplicates
if SHEET_NAME in wb.sheetnames:
    wb.remove(wb[SHEET_NAME])
# Create new sheet at index 0
ws = wb.create_sheet(title=SHEET_NAME, index=0)
# Write header and content lines
ws.append(['content'])
for ln in selected:
    ws.append([ln])

wb.save(XLSX_PATH)
print(f'Wrote {len(selected)} lines to sheet "{SHEET_NAME}" in {XLSX_PATH}')
