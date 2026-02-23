#!/usr/bin/env python3
from pathlib import Path
import re
from openpyxl import Workbook

MD_PATH = Path('Doc/Testing/system_test_specification.md')
XLSX_PATH = Path('Doc/Testing/system_test_specification_tables.xlsx')

# Settings
SELECT_LINES = 67  # first 1..67 lines to '選択範囲'
SECTION_KEYS = {
    '3.1': '3.1 一般ユーザーシナリオ',
    '3.2': '3.2 管理ユーザーシナリオ',
    '3.3': '3.3 エラーハンドリングシナリオ',
}

FORBIDDEN_SHEET = r'[:\\/?*\[\]]'


def sanitize_sheet_name(name: str) -> str:
    name = re.sub(FORBIDDEN_SHEET, '_', name)
    return name.strip()[:31]


def split_row(r: str):
    r = r.strip()
    if r.startswith('|') and r.endswith('|'):
        r = r[1:-1]
    parts = [c.strip() for c in r.split('|')]
    return parts


def convert_tables_to_tabs(lines):
    out = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if '|' in line and i+1 < len(lines) and re.search(r'^[\s\|:\-]+$', lines[i+1].replace(' ', '')):
            # table block
            header = lines[i]
            j = i+2
            rows = []
            while j < len(lines) and '|' in lines[j]:
                if re.match(r'^\s*$', lines[j]):
                    break
                rows.append(lines[j])
                j += 1
            # convert header and rows
            def to_tab(r):
                parts = split_row(r)
                return '\t'.join(parts)
            out.append(to_tab(header))
            for r in rows:
                out.append(to_tab(r))
            i = j
            continue
        else:
            out.append(line)
            i += 1
    return out


def parse_for_workbook(lines):
    # returns selection_lines, tables_by_subsection (3.1/3.2/3.3), individual_tests {sheet_name: rows}
    selection = lines[:SELECT_LINES]
    tables_by_sub = {k: [] for k in SECTION_KEYS.keys()}
    individual = {}

    current_sub = None
    current_test = None

    i = 0
    while i < len(lines):
        line = lines[i]
        m = re.match(r'^(#{1,6})\s*(.*)$', line)
        if m:
            heading = m.group(2).strip()
            # check subsection
            for key in SECTION_KEYS.keys():
                if heading.startswith(key):
                    current_sub = key
                    break
            else:
                # if heading is a test case like UST-001 or AST-001 or EST-001
                tm = re.match(r'^(UST|AST|EST)-\d{3}[:_\s-]*(.*)$', heading)
                if tm:
                    code = tm.group(0).split()[0] if tm.group(0) else None
                    # code full e.g. 'UST-001' get short name
                    code_match = re.match(r'^(UST|AST|EST)-\d{3}', heading)
                    if code_match:
                        code_id = code_match.group(0)
                        title = heading[len(code_id):].strip(' :-_')
                        sheet_name = f"{code_id}_{title}" if title else code_id
                        sheet_name = sanitize_sheet_name(sheet_name)
                        current_test = sheet_name
                        # initialize individual slot
                        individual[current_test] = []
                    else:
                        current_test = None
                else:
                    current_test = None
            i += 1
            continue

        # detect table
        if '|' in line and i+1 < len(lines) and re.search(r'^[\s\|:\-]+$', lines[i+1].replace(' ', '')):
            header_line = line
            j = i+2
            rows = []
            while j < len(lines) and '|' in lines[j]:
                if re.match(r'^\s*$', lines[j]):
                    break
                rows.append(lines[j])
                j += 1
            header = split_row(header_line)
            data = [split_row(r) for r in rows if r.strip()]
            if current_test is not None:
                # store in individual test first table found (append)
                individual[current_test].append((header, data))
            elif current_sub in tables_by_sub:
                tables_by_sub[current_sub].append((header, data))
            # else ignore
            i = j
            continue
        i += 1

    return selection, tables_by_sub, individual


def create_workbook(selection, tables_by_sub, individual, out_path: Path):
    wb = Workbook()
    # remove default
    wb.remove(wb.active)

    # 1) 選択範囲 sheet
    ws_sel = wb.create_sheet(title='選択範囲')
    ws_sel.append(['content'])
    for ln in selection:
        ws_sel.append([ln])

    # 2) sheets for 3.1,3.2,3.3
    for key, title in SECTION_KEYS.items():
        ws = wb.create_sheet(title=title[:31])
        ws.append([title])
        ws.append([''])
        tables = tables_by_sub.get(key, [])
        if not tables:
            ws.append(['(No tables found in this section)'])
            continue
        for idx, (header, rows) in enumerate(tables, start=1):
            ws.append([f'Table {idx}'])
            ws.append(header)
            for r in rows:
                ws.append(r)
            ws.append([''])

    # 3) individual test case sheets
    for sheet, tables in individual.items():
        ws = wb.create_sheet(title=sheet)
        if not tables:
            ws.append(['(No tables found)'])
            continue
        for idx, (header, rows) in enumerate(tables, start=1):
            ws.append([f'Table {idx}'])
            ws.append(header)
            for r in rows:
                ws.append(r)
            ws.append([''])

    wb.save(out_path)
    print(f'Created workbook: {out_path} (選択範囲 + 3.1/3.2/3.3 + {len(individual)} individual test sheets)')


if __name__ == '__main__':
    import sys
    if not MD_PATH.exists():
        print('MD not found:', MD_PATH); raise SystemExit(1)
    text = MD_PATH.read_text(encoding='utf-8')
    lines = text.splitlines()

    # 1) convert tables in md to tab-separated and write back
    converted = convert_tables_to_tabs(lines)
    MD_PATH.write_text('\n'.join(converted), encoding='utf-8')
    print('Converted markdown tables to tabs in', MD_PATH)

    # 2) parse for workbook
    selection, tables_by_sub, individual = parse_for_workbook(converted)

    # 3) create workbook
    create_workbook(selection, tables_by_sub, individual, XLSX_PATH)
